#!/usr/bin/env python3

# tkauto.py
# Python console program
# Author: Michael Leidel
# Description:
# Builds a Python tkinter application shell from an xlsx file (layout.xlsx)
# with these columns:
# Widget, Variable, Text, Command/Textvariable, Row, Column, Rowspan, Colspan, Sticky

import openpyxl
import csv, os
import argparse

parser = argparse.ArgumentParser(description='Tkauto build python GUI')
parser.add_argument('filename', help='Excel (xlsx) file to use as input')
args = parser.parse_args()

if os.path.exists(args.filename):
    if not args.filename.endswith("xlsx"):
        print("must be an Excel (xlsx) file.")
        exit()
else:
    print("cannot find file: " + args.filename)
    exit()

# list (flds) index names for the (row)column values
nop, wgt, var, txt, com, row, col, rsp, csp, sty = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9

sout = ""  # holds output for injection into template


def prt(s):
    ''' Adjust leading tabs/spaces for output here '''
    global sout
    # print("        " + s)
    sout += "        " + s + "\n"


def getRowSpan(val):
    rowspan = ""
    if val != "":
        rowspan = ", rowspan=" + str(val)
    return rowspan


def getColSpan(val):
    colspan = ""
    if val != "":
        colspan = ", columnspan=" + str(val)
    return colspan


def getSticky(val):
    sticky = ""
    if val != "":
        sticky = ", sticky=" + val
    return sticky


'''
    Get the Excel workbook and spreadsheet.
    The Excel file is expected to be in the app's directory.
'''

wb = openpyxl.load_workbook(args.filename)
sheet = wb.get_sheet_by_name('layout')  # Must be a Sheet titled 'layout' !
flds = []
callbacks = []

for rownum in range(2, sheet.max_row + 1):  # loop through each row

    flds.clear()  # clear list
    flds.append("nop")  # zero element not used
    # load up the flds list with this row's columns values
    for c in range(1, 10):  # columns align with list index
        val = sheet.cell(row=rownum, column=c).value
        if val == None:
            flds.append("")
        else:
            flds.append(val)

    # Process this row/columns values

    # BUTTON
    if flds[wgt].lower() == "button":
        line = "{0} = Button(self, text='{1}', command=self.{2})"
        prt(line.format(flds[var], flds[txt], flds[com]))
        callbacks.append(flds[com])
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # LABEL
    elif flds[wgt].lower() == "label":
        prt("self." + flds[com] + " = StringVar()")
        line = "{} = Label(self, text='{}', textvariable=self.{})"
        prt(line.format(flds[var], flds[txt], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # ENTRY
    elif flds[wgt].lower() == "entry":
        prt("self." + flds[com] + " = StringVar()")
        prt("# self." + flds[com] + ".trace(\"w\", self.eventHandler)")
        line = "{0} = Entry(self, textvariable=self.{1})"
        prt(line.format(flds[var], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # TEXT
    elif flds[wgt].lower() == "text":
        line = "self.{0} = Text(self, relief=SUNKEN)"
        prt(line.format(flds[var]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "self.{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''
        # efont = Font(family="Monospace", size=14)
        # self.EDITOR.configure(font=efont)
        # self.EDITOR.config(wrap = "word", # wrap = NONE
        #        undo = True, # Tk 8.4
        #        width = 80,
        #        tabs = (efont.measure(' ' * 4),))
        # self.EDITOR.focus()
        ## basic handler commands #
        # .get("1.0", END)
        # .delete("1.0", END)
        # .insert("1.0", "New text content ...")
        '''
        prt(line + "\n")

    # LIST
    elif flds[wgt].lower() == "list":
        line = "self.{0} = Listbox(self, height=5)"
        prt(line.format(flds[var]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "self.{0}.grid(row={1}, column={2} {3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = '''
        # self.LISTBOX.bind("<Double-Button-1>", self.open_path)
        # for i in range(100):
        #     self.LISTBOX.insert(i, str(i) + "Item")'''
        prt(line)
        line = '''
    ## Handler for List selection
    ## Make this a class method
    # def open_path(self, event):
    #     list_item = self.LISTBOX.curselection()
    #     fp = self.LISTBOX.get(list_item[0])
    #     print(str(fp) + " --> " + str(list_item[0]) +
    #         " of " + str(self.LISTBOX.size()))
    #
    # FUNCS TO EDIT LISTBOX CONTENTS
    #
    # def delete_item(self):
    #     if self.listbox.curselection() == ():
    #         return # nothing selected
    #     print("Deleting: " + str(self.listbox.curselection()))
    #     self.listbox.delete(self.listbox.curselection())

    # def insert_item(self):
    #     if self.listbox.curselection() == ():
    #         return # nothing selected
    #     list_item = self.listbox.curselection()
    #     self.listbox.insert(list_item[0], self.txtfld.get())
    #     print("inserted at " + str(list_item[0]))
        '''
        prt(line + "\n")

    # VERT SCROLLBAR
    elif flds[wgt].lower() == "scrolly":
        line = "self.{0} = Scrollbar(self, orient=VERTICAL, command=self.{1}.yview)"
        if flds[com] == "":
            print("\n\nMISSING list object FOR SCROLLBAR WIDGET\n\n")
            exit()
        prt(line.format(flds[var], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "self.{0}.grid(row={1}, column={2} {3}{4}{5})  # use N+S+E"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = "self.{0}['yscrollcommand'] = self.{1}.set\n"
        prt(line.format(flds[com], flds[var]))

    # HORZ SCROLLBAR
    elif flds[wgt].lower() == "scrollx":
        line = "self.{0} = Scrollbar(self, orient=HORIZONTAL, command=self.{1}.xview)"
        if flds[com] == "":
            print("\n\nMISSING list object FOR SCROLLBAR WIDGET\n\n")
            exit()
        prt(line.format(flds[var], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "self.{0}.grid(row={1}, column={2} {3}{4}{5})"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
        line = "self.{0}['xscrollcommand'] = self.{1}.set\n"
        prt(line.format(flds[com], flds[var]))

    # CHECK BOX
    elif flds[wgt].lower() == "check":
        prt("self.{0} = IntVar()".format(flds[com]))
        line = "{0} = Checkbutton(self, relief=RAISED, variable=self.{1}, text='{2}')"
        prt(line.format(flds[var], flds[com], flds[txt]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # RADIO BUTTON
    elif flds[wgt].lower() == "radio":
        prt("self." + flds[com] +
            " = StringVar() # Use one Var per group of buttons")
        line = "{0} = Radiobutton(self, variable=self.{1}, value='{2}', text='{2}')"
        prt(line.format(flds[var], flds[com], flds[txt], flds[txt]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # SPIN BOX
    elif flds[wgt].lower() == "spin":
        prt("self." + flds[com] + " = StringVar()")
        line = "{0} = Spinbox(self, from_=1, to=9, relief=SUNKEN, width=2, textvariable=self.{1})"
        prt(line.format(flds[var], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))
    # OPTION MENU
    elif flds[wgt].lower() == "options":
        prt("optionlist = ('aaa', 'bbb', 'ccc', 'ddd', 'eee', 'fff')")
        prt("self." + flds[com] + " = StringVar()")
        prt("self." + flds[com] + ".set(optionlist[0])")
        line = "{0} = OptionMenu(self, self.{1}, *optionlist)"
        prt(line.format(flds[var], flds[com]))
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    # COMBOBOX (from tkinter.ttk import Combobox)
    elif flds[wgt].lower() == "combo":
        prt("self." + flds[com] + " = StringVar()")
        line = "{0} = Combobox(self, textvariable=self.{1})"
        prt(line.format(flds[var], flds[com]))
        line = "{0}['values'] = ('value1', 'value2', 'value3')"
        prt(line.format(flds[var]))
        prt("# COMBO.bind('<<ComboboxSelected>>', self.ONCOMBOSELECT)")
        prt("%s.current(0)" % flds[var])
        rowspan = getRowSpan(flds[rsp])
        colspan = getColSpan(flds[csp])
        sticky = getSticky(flds[sty])
        line = "{0}.grid(row={1}, column={2} {3}{4}{5})\n"
        prt(line.format(flds[var], flds[row],
                        flds[col], rowspan, colspan, sticky))

    elif flds[wgt].lower() == "frames":
        line = '''
        # lframe = LabelFrame(self, text="text",
        #                     width=100, height=100)
        # lframe.grid(row=1, column=1, sticky=N+S+E+W)
        #
        # fram = Frame(self, width=100, height=100)
        # fram.grid(row=1, column=1, sticky=N+S+E+W)
        '''
        prt(line + "\n")

    elif flds[wgt].lower() == "messagebox":
        line = '''
        # from tkinter import messagebox
        # messagebox.showerror("Error", "Error message")
        # messagebox.showwarning("Warning","Warning message")
        # messagebox.showinfo("Information","Informative message")
        # messagebox.askokcancel('Message title', 'Message content')
        # messagebox.askretrycancel('Message title', 'Message content')
        #     ok, yes, retry returns TRUE
        #     no, cancel returns FALSE
        '''
        prt(line + "\n")

    elif flds[wgt].lower() == "filedialog":
        line = '''
        # from tkinter import filedialog
        # filename =  filedialog.askopenfilename(initialdir = "/",
        #             title = "Open file",
        #             filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
        # filename = filedialog.asksaveasfilename(initialdir = "/",
        #             title = "Save file",
        #             filetypes = (("jpeg files","*.jpg"),("all files","*.*")))'''
        prt(line + "\n")

    elif flds[wgt].lower() == "geometry":
        line = "root.geometry(\"%s\")" % (flds[var])
        prt(line + "\n")

    else:
        if flds[wgt].startswith("Widget"):
            pass
        else:
            prt("\nNUNRECOGNIZED WIDGET IDENTIFIER: " + widget + "\n\n")

fout = open("output.py", "w")
fin = open("/home/ml/apps/python/projects/tkautox/tkauto_tpl.py", "r")
for line in fin:
    if line.find("INSERT TKAUTO OUTPUT") > 0:
        fout.write(sout)
        #  insert the callbacks
        for item in callbacks:
            fout.write("    def %s(self):\n" % (item))
            fout.write("        ''' docstring '''\n\n")
    else:
        fout.write(line)


fout.close()
fin.close()
print("\n\nFind new script in 'output.py'\n")
print("Some Widgets may need futher definition\n")
